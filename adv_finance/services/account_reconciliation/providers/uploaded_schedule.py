from __future__ import annotations

import csv
from decimal import Decimal
from pathlib import Path

import frappe
from frappe.utils.file_manager import get_file_path

from adv_finance.services.account_reconciliation.providers.base import ReconciliationProvider
from adv_finance.services.statement_normalizer import normalize_decimal


class UploadedScheduleProvider(ReconciliationProvider):
    provider_name = "Uploaded Schedule"

    def validate(self, reconciliation) -> None:
        if not reconciliation.supporting_document:
            frappe.throw("Supporting document is required for Uploaded Schedule reconciliations.")

    def get_supporting_balance(self, reconciliation):
        self.validate(reconciliation)
        path = Path(get_file_path(reconciliation.supporting_document))
        if path.suffix.lower() == ".xlsx":
            return _read_xlsx_amount_total(path)
        if path.suffix.lower() != ".csv":
            frappe.throw("Uploaded schedule provider supports CSV and XLSX schedules.")
        return _read_csv_amount_total(path)


def _read_csv_amount_total(path: Path) -> Decimal:
    total = Decimal("0")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if "amount" not in [field.lower() for field in (reader.fieldnames or [])]:
            frappe.throw('Uploaded schedule must include an "amount" column.')
        amount_field = next(field for field in reader.fieldnames if field.lower() == "amount")
        for row in reader:
            total += normalize_decimal(row.get(amount_field))
    return total


def _read_xlsx_amount_total(path: Path) -> Decimal:
    try:
        from openpyxl import load_workbook
    except ImportError:
        frappe.throw("XLSX schedule parsing requires openpyxl in the Frappe environment.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return Decimal("0")
    headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
    if "amount" not in headers:
        frappe.throw('Uploaded schedule must include an "amount" column.')
    amount_index = headers.index("amount")
    total = Decimal("0")
    for row in rows[1:]:
        if amount_index < len(row):
            total += normalize_decimal(row[amount_index])
    return total
