from __future__ import annotations

import csv
import mimetypes
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

import frappe
from frappe.utils.file_manager import get_file_path

from adv_finance.services.statement_normalizer import (
    classify_transaction_type,
    normalize_date,
    normalize_decimal,
    normalize_description,
    normalize_reference,
)


MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024


@dataclass(frozen=True)
class ParsedRow:
    line_number: int
    values: dict[str, Any]


def parse_supplier_statement(reconciliation_name: str, force: bool = False) -> dict:
    reconciliation = frappe.get_doc("Supplier Reconciliation", reconciliation_name)
    _ensure_can_rebuild(reconciliation, force)

    if not reconciliation.statement_file:
        frappe.throw("Attach a supplier statement before parsing.")
    if not reconciliation.statement_template:
        frappe.throw("Select a Supplier Statement Template before parsing.")

    template = frappe.get_doc("Supplier Statement Template", reconciliation.statement_template)
    file_path = Path(get_file_path(reconciliation.statement_file))
    _validate_file(file_path, template.file_type)

    rows = _read_rows(file_path, template)
    lines = [_build_line(row, template) for row in rows]

    reconciliation.set("statement_lines", [])
    for line in lines:
        reconciliation.append("statement_lines", line)

    closing_balance = _statement_closing_balance(lines, reconciliation.statement_closing_balance)
    reconciliation.update(
        {
            "statement_closing_balance": closing_balance,
            "total_statement_lines": len(lines),
            "reconciliation_status": "Parsed",
        }
    )
    reconciliation.save()

    return {"reconciliation": reconciliation.name, "statement_lines": len(lines)}


def _ensure_can_rebuild(reconciliation, force: bool):
    if reconciliation.reconciliation_status == "Closed":
        frappe.throw("Closed reconciliations cannot be reparsed. Reopen first.")
    if reconciliation.statement_lines and not force:
        frappe.throw("Statement lines already exist. Re-run with force=True to rebuild them.")


def _validate_file(file_path: Path, file_type: str):
    if not file_path.exists():
        frappe.throw("Attached statement file was not found on the server.")
    if file_path.stat().st_size == 0:
        frappe.throw("Attached statement file is empty.")
    if file_path.stat().st_size > MAX_FILE_SIZE_BYTES:
        frappe.throw("Attached statement file exceeds the 10 MB limit.")

    extension = file_path.suffix.lower()
    expected = ".csv" if file_type == "CSV" else ".xlsx"
    if extension != expected:
        frappe.throw(f"Statement template expects a {expected} file.")

    mime_type, _encoding = mimetypes.guess_type(str(file_path))
    if file_type == "CSV" and mime_type not in (None, "text/csv", "text/plain", "application/vnd.ms-excel"):
        frappe.throw("Attached file does not look like a CSV file.")


def _read_rows(file_path: Path, template) -> list[ParsedRow]:
    if template.file_type == "CSV":
        return _read_csv_rows(file_path, template)
    return _read_xlsx_rows(file_path, template)


def _read_csv_rows(file_path: Path, template) -> list[ParsedRow]:
    header_row = int(template.header_row_number or 1)
    with file_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = list(csv.reader(handle))

    if len(reader) < header_row:
        frappe.throw(f"Statement file does not contain header row {header_row}.")

    headers = [str(value).strip() for value in reader[header_row - 1]]
    result = []
    for index, row in enumerate(reader[header_row:], start=header_row + 1):
        if template.skip_blank_rows and not any(row):
            continue
        result.append(ParsedRow(index, dict(zip(headers, row, strict=False))))
    return result


def _read_xlsx_rows(file_path: Path, template) -> list[ParsedRow]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        frappe.throw("XLSX parsing requires openpyxl to be available in the Frappe environment.")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet_name = template.sheet_name or workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        frappe.throw(f'Sheet "{sheet_name}" was not found in the uploaded workbook.')

    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    header_row = int(template.header_row_number or 1)
    if len(rows) < header_row:
        frappe.throw(f"Workbook does not contain header row {header_row}.")

    headers = [str(value).strip() if value is not None else "" for value in rows[header_row - 1]]
    result = []
    for index, row in enumerate(rows[header_row:], start=header_row + 1):
        if template.skip_blank_rows and not any(row):
            continue
        result.append(ParsedRow(index, dict(zip(headers, row, strict=False))))
    return result


def _build_line(row: ParsedRow, template) -> dict[str, Any]:
    def mapped(column_field: str):
        column_name = template.get(column_field)
        if not column_name:
            return None
        if column_name not in row.values:
            frappe.throw(f'Statement template expects column "{column_name}", but it was not found in row {row.line_number}.')
        return row.values.get(column_name)

    raw_date = mapped("transaction_date_column")
    raw_reference = mapped("reference_column")
    raw_description = mapped("description_column")
    raw_debit = mapped("debit_column")
    raw_credit = mapped("credit_column")
    raw_amount = mapped("amount_column")
    raw_balance = mapped("balance_column")

    debit = normalize_decimal(raw_debit, template.decimal_separator, template.thousands_separator)
    credit = normalize_decimal(raw_credit, template.decimal_separator, template.thousands_separator)
    amount = normalize_decimal(raw_amount, template.decimal_separator, template.thousands_separator) if raw_amount not in (None, "") else debit - credit
    balance = normalize_decimal(raw_balance, template.decimal_separator, template.thousands_separator)
    description = normalize_description(raw_description)

    return {
        "line_number": row.line_number,
        "transaction_date": normalize_date(raw_date, template.date_format),
        "posting_date": normalize_date(raw_date, template.date_format),
        "transaction_type": mapped("transaction_type_column") or classify_transaction_type(description, debit, credit),
        "reference": raw_reference,
        "normalized_reference": normalize_reference(raw_reference),
        "secondary_reference": mapped("secondary_reference_column"),
        "description": description,
        "debit": debit,
        "credit": credit,
        "amount": amount,
        "running_balance": balance,
        "currency": None,
        "match_status": "Unmatched",
        "raw_date": raw_date,
        "raw_reference": raw_reference,
        "raw_description": raw_description,
        "raw_debit": raw_debit,
        "raw_credit": raw_credit,
        "raw_amount": raw_amount,
        "raw_balance": raw_balance,
    }


def _statement_closing_balance(lines: list[dict[str, Any]], manual_balance) -> Decimal:
    if lines:
        last_balance = lines[-1].get("running_balance")
        if last_balance is not None:
            return last_balance
    return Decimal(str(manual_balance or 0))
